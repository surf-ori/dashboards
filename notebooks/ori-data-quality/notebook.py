# /// script
# requires-python = ">=3.12"
# dependencies = [
#     "altair==6.1.0",
#     "duckdb==1.5.2",
#     "jedi<0.20.0",
#     "marimo>=0.10.0",
#     "openpyxl==3.1.5",
#     "pandas==3.0.2",
#     "polars[pyarrow]==1.39.3",
#     "pyarrow==24.0.0",
#     "sqlglot==30.6.0",
# ]
# ///

import marimo

__generated_with = "0.23.5"
app = marimo.App(width="full", app_title="ORI Data Quality Dashboard")


@app.cell(hide_code=True)
async def wasm_dependencies():
    # install packages not bundled by the WASM export (duckdb is handled by mo.sql)
    import sys
    import io
    import openpyxl
    _ = None
    if 'pyodide' in sys.modules:
        import micropip
        await micropip.install(['polars', 'pyarrow', 'altair', 'openpyxl'])
        _ = 'wasm'
    return io, openpyxl, sys


@app.cell(hide_code=True)
def imports():
    # import all third-party libraries used throughout the notebook
    import marimo as mo
    import polars as pl
    import altair as alt

    return alt, mo, pl


@app.cell(hide_code=True)
def cache_setup(mo, sys):
    from datetime import datetime
    import hashlib

    is_wasm = 'pyodide' in sys.modules
    CACHE_DIR = mo.notebook_location() / 'cache'

    _sentinel = CACHE_DIR / '.last_refreshed'
    if not is_wasm and _sentinel.exists():
        _ts = datetime.fromtimestamp(_sentinel.stat().st_mtime)
        cache_last_refreshed = _ts.strftime('%Y-%m-%d %H:%M')
    else:
        cache_last_refreshed = None
    return CACHE_DIR, cache_last_refreshed, hashlib, is_wasm


@app.cell(hide_code=True)
def refresh_ui(mo):
    refresh_btn = mo.ui.button(
        value=0,
        on_click=lambda v: v + 1,
        label=f"{mo.icon('lucide:refresh-cw')} Refresh all data",
        kind='warn',
        full_width=True,
    )
    return (refresh_btn,)


@app.cell(hide_code=True)
def catalog_url(mo):
    # text input widget for the ORI DuckLake catalog URL
    url = mo.ui.text(
        value='https://objectstore.surf.nl/cea01a7216d64348b7e51e5f3fc1901d:sprouts/catalog.ducklake',
        full_width=True,
    )
    mo.accordion({"DuckLake catalog URL": url})
    return (url,)


@app.cell(hide_code=True)
def attach_catalog(mo, url):
    # attach the ORI DuckLake catalog to the shared mo.sql DuckDB connection
    mo.sql(
        f"""
        ATTACH '{url.value}' AS sprouts (TYPE ducklake, READ_ONLY);
        USE sprouts;
        """
    )
    return


@app.cell(hide_code=True)
async def load_nl_baseline(
    CACHE_DIR,
    io,
    is_wasm,
    openpyxl,
    pl,
    refresh_btn,
    sys,
):
    # fetch the Dutch research organisations baseline list from Zenodo (DOI: 10.5281/zenodo.18957154)
    # and annotate each org with its Barcelona Declaration signatory status
    _force = refresh_btn.value > 0
    _cache = CACHE_DIR / 'nl_baseline.parquet'
    if not is_wasm and not _force and _cache.exists():
        nl_baseline_df = pl.read_parquet(_cache)
    else:
        _ZENODO_URL = 'https://zenodo.org/api/records/18957154/files/nl-orgs-baseline.xlsx/content'

        # ROR IDs of NL Barcelona Declaration signatories (source: barcelona-declaration.org/signatories_by_country/)
        _BARCELONA_RORS = {
            'https://ror.org/02e2c7k09',  # Delft University of Technology
            'https://ror.org/04jsz6e67',  # Dutch Research Council (NWO)
            'https://ror.org/02w4jbg70',  # KB, National Library of the Netherlands
            'https://ror.org/027bh9e22',  # Leiden University
            'https://ror.org/00rbjv475',  # Netherlands eScience Center
            'https://ror.org/043c0p156',  # Royal Netherlands Academy of Arts and Sciences (KNAW)
            'https://ror.org/009vhk114',  # SURF
            'https://ror.org/04dkp9463',  # University of Amsterdam
            'https://ror.org/012p63287',  # University of Groningen
            'https://ror.org/04pp8hn57',  # Utrecht University
            'https://ror.org/008xxew50',  # Vrije Universiteit Amsterdam
        }

        if 'pyodide' in sys.modules:
            import pyodide.http
            _resp = await pyodide.http.pyfetch(_ZENODO_URL)
            _raw = await _resp.bytes()
        else:
            import urllib.request
            with urllib.request.urlopen(_ZENODO_URL) as _r:
                _raw = _r.read()

        _wb = openpyxl.load_workbook(io.BytesIO(_raw), read_only=True)
        _ws = _wb['nl-orgs']
        _all_rows = list(_ws.iter_rows(values_only=True))
        _data_rows = [r for r in _all_rows[1:] if r[0] and r[3]]  # skip header + empty rows

        nl_baseline_df = pl.DataFrame({
            'full_name': [r[0] for r in _data_rows],
            'acronym':   [r[1] or '' for r in _data_rows],
            'grouping':  [r[3] for r in _data_rows],
            'ror':       [r[5] if r[5] and r[5] != 'nvt' else None for r in _data_rows],
        }).with_columns(
            pl.col('ror').is_in(_BARCELONA_RORS).fill_null(False).alias('barcelona_signatory'),
        )
        if not is_wasm:
            CACHE_DIR.mkdir(parents=True, exist_ok=True)
            nl_baseline_df.write_parquet(_cache)
            (CACHE_DIR / '.last_refreshed').touch()
    return (nl_baseline_df,)


@app.cell(hide_code=True)
def load_nl_openalex_orgs(
    CACHE_DIR,
    is_wasm,
    mo,
    nl_baseline_df,
    pl,
    refresh_btn,
):
    # query OpenAlex institutions for orgs in the NL baseline, anchored by ROR
    _force = refresh_btn.value > 0
    _cache = CACHE_DIR / 'nl_openalex_orgs.parquet'
    if not is_wasm and not _force and _cache.exists():
        nl_openalex_orgs_df = pl.read_parquet(_cache)
    else:
        _rors = nl_baseline_df['ror'].drop_nulls().to_list()
        _rors_clause = ', '.join(f"'{r}'" for r in _rors) if _rors else "''"
        nl_openalex_orgs_df = mo.sql(
            f"""
            SELECT
                id AS openalex_orgs_id,
                display_name,
                ror,
                type,
                works_count,
                cited_by_count,
                ids.ror      IS NOT NULL AS has_ror,
                ids.grid     IS NOT NULL AS has_grid,
                ids.wikidata IS NOT NULL AS has_wikidata,
                ids.wikipedia IS NOT NULL AS has_wikipedia,
                homepage_url IS NOT NULL AS has_homepage,
                ids.grid      AS grid_id,
                ids.wikidata  AS wikidata_id,
                ids.wikipedia AS wikipedia_url,
                homepage_url
            FROM openalex.institutions
            WHERE ror IN ({_rors_clause})
            ORDER BY works_count DESC
            """,
            output=False
        )
        if not is_wasm:
            CACHE_DIR.mkdir(parents=True, exist_ok=True)
            nl_openalex_orgs_df.write_parquet(_cache)
            (CACHE_DIR / '.last_refreshed').touch()
    return (nl_openalex_orgs_df,)


@app.cell(hide_code=True)
def load_nl_openaire_orgs(
    CACHE_DIR,
    is_wasm,
    mo,
    nl_baseline_df,
    pl,
    refresh_btn,
):
    # query OpenAIRE organizations for orgs in the NL baseline, anchored by ROR
    _force = refresh_btn.value > 0
    _cache = CACHE_DIR / 'nl_openaire_orgs.parquet'
    if not is_wasm and not _force and _cache.exists():
        nl_openaire_orgs_df = pl.read_parquet(_cache)
    else:
        _rors = nl_baseline_df['ror'].drop_nulls().to_list()
        _rors_list = '[' + ', '.join(f"'{r}'" for r in _rors) + ']' if _rors else "['']"
        nl_openaire_orgs_df = mo.sql(f"""
        SELECT
            o.legalName,
            o.legalShortName,
            o.websiteUrl,
            o.id AS openaire_orgs_id,
            o.pids
        FROM openaire.organizations AS o,
             UNNEST(o.pids) AS unnest
        WHERE unnest.scheme = 'ROR'
          AND list_contains({_rors_list}, unnest.value)
        """, output=False)
        # Extract PIDs from pids list
        nl_openaire_orgs_df = nl_openaire_orgs_df.with_columns(
            pl.col('pids').list.eval(
                pl.when(pl.element().struct.field('scheme').str.to_uppercase() == 'ROR')
                .then(pl.element().struct.field('value'))
            ).list.drop_nulls().list.first().alias('ror'),
            pl.col('pids').list.eval(
                pl.when(pl.element().struct.field('scheme').str.to_uppercase() == 'ISNI')
                .then(pl.element().struct.field('value'))
            ).list.drop_nulls().list.first().alias('isni'),
            pl.col('pids').list.eval(
                pl.when(pl.element().struct.field('scheme').str.to_uppercase() == 'GRID')
                .then(pl.element().struct.field('value'))
            ).list.drop_nulls().list.first().alias('grid'),
            pl.col('pids').list.eval(
                pl.when(pl.element().struct.field('scheme').str.to_uppercase() == 'WIKIDATA')
                .then(pl.element().struct.field('value'))
            ).list.drop_nulls().list.first().alias('wikidata'),
        )
        if not is_wasm:
            CACHE_DIR.mkdir(parents=True, exist_ok=True)
            nl_openaire_orgs_df.write_parquet(_cache)
            (CACHE_DIR / '.last_refreshed').touch()
    return (nl_openaire_orgs_df,)


@app.cell(hide_code=True)
async def load_nl_endpoint_table(
    CACHE_DIR,
    io,
    is_wasm,
    openpyxl,
    pl,
    refresh_btn,
    sys,
):
    # fetch the NL organisations → OpenAIRE datasource endpoint table from Zenodo
    # (DOI: 10.5281/zenodo.18959652); maps openaire_org_id to OpenAIRE_DataSource_ID + OAI endpoints
    _force = refresh_btn.value > 0
    _cache = CACHE_DIR / 'nl_endpoint.parquet'
    if not is_wasm and not _force and _cache.exists():
        nl_endpoint_df = pl.read_parquet(_cache)
    else:
        _ZENODO_URL = 'https://zenodo.org/api/records/19470205/files/nl_orgs_openaire_datasources_with_endpoint_public.xlsx/content'

        if 'pyodide' in sys.modules:
            import pyodide.http as _pyodide_http
            _resp = await _pyodide_http.pyfetch(_ZENODO_URL)
            _raw = await _resp.bytes()
        else:
            import urllib.request as _urllib_request
            with _urllib_request.urlopen(_ZENODO_URL) as _r:
                _raw = _r.read()

        _wb = openpyxl.load_workbook(io.BytesIO(_raw), read_only=True)
        _ws = _wb.active
        _all_rows = list(_ws.iter_rows(values_only=True))
        _headers = [str(h) if h is not None else f'col_{i}' for i, h in enumerate(_all_rows[0])]
        nl_endpoint_df = pl.DataFrame(
            {_headers[i]: [r[i] for r in _all_rows[1:]] for i in range(len(_headers))}
        )
        if not is_wasm:
            CACHE_DIR.mkdir(parents=True, exist_ok=True)
            nl_endpoint_df.write_parquet(_cache)
            (CACHE_DIR / '.last_refreshed').touch()
    return (nl_endpoint_df,)


@app.cell(hide_code=True)
def load_nl_openaire_datasources(
    CACHE_DIR,
    is_wasm,
    mo,
    nl_endpoint_df,
    pl,
    refresh_btn,
):
    # query OpenAIRE datasources table for datasource IDs from the endpoint table
    _force = refresh_btn.value > 0
    _cache = CACHE_DIR / 'nl_openaire_datasources.parquet'
    if not is_wasm and not _force and _cache.exists():
        nl_datasources_df = pl.read_parquet(_cache)
    else:
        _ds_ids = nl_endpoint_df['OpenAIRE_DataSource_ID'].drop_nulls().unique().to_list()
        _ids_clause = ', '.join(f"'{i}'" for i in _ds_ids) if _ds_ids else "''"
        nl_datasources_df = mo.sql(f"""
        SELECT *
        FROM openaire.datasources
        WHERE id IN ({_ids_clause})
        """, output=False)
        if not is_wasm:
            CACHE_DIR.mkdir(parents=True, exist_ok=True)
            nl_datasources_df.write_parquet(_cache)
            (CACHE_DIR / '.last_refreshed').touch()
    return (nl_datasources_df,)


@app.cell(hide_code=True)
def load_openalex_publications_counts(
    CACHE_DIR,
    hashlib,
    is_wasm,
    mo,
    nl_baseline_df,
    nl_openalex_orgs_df,
    org_select,
    pl,
    refresh_btn,
):
    # Count OpenAlex works for selected organisations by querying openalex.works directly.
    # Uses openalex_orgs_id (institution ID) to filter authorships[].institutions[].id.
    # WARNING: full scan of openalex.works (364 M rows via UNNEST) — expect 15+ minutes.
    _sel_rors = (
        nl_baseline_df
        .filter(pl.col('full_name').is_in(org_select.value))
        ['ror'].drop_nulls().to_list()
    )
    _openalex_ids = (
        nl_openalex_orgs_df
        .filter(pl.col('ror').is_in(_sel_rors))
        ['openalex_orgs_id'].drop_nulls().to_list()
    )
    _force = refresh_btn.value > 0
    _key = hashlib.sha1(','.join(sorted(_sel_rors)).encode()).hexdigest()[:8]
    _cache = CACHE_DIR / f'openalex_works_{_key}.parquet'
    if not is_wasm and not _force and _cache.exists():
        openalex_works_df = pl.read_parquet(_cache)
    else:
        _openalex_ids_list = '[' + ', '.join(f"'{i}'" for i in _openalex_ids) + ']' if _openalex_ids else "['']"
        openalex_works_df = mo.sql(f"""
        SELECT COUNT(DISTINCT w.id) AS openalex_works_count
        FROM openalex.works AS w,
             UNNEST(w.authorships) AS unnest
        WHERE array_length(list_filter(
            unnest.institutions,
            x -> list_contains({_openalex_ids_list}, x.id)
        )) > 0
        """, output=False)
        if not is_wasm:
            CACHE_DIR.mkdir(parents=True, exist_ok=True)
            openalex_works_df.write_parquet(_cache)
            (CACHE_DIR / '.last_refreshed').touch()
    return (openalex_works_df,)


@app.cell(hide_code=True)
def load_openaire_pubs_counts(
    CACHE_DIR,
    hashlib,
    is_wasm,
    mo,
    nl_baseline_df,
    nl_openaire_orgs_df,
    org_select,
    pl,
    refresh_btn,
):
    # Count OpenAIRE publications for selected organisations using openaire_orgs_id.
    # Filters publications.organizations[].id — full scan of 206 M rows.
    _sel_rors = (
        nl_baseline_df
        .filter(pl.col('full_name').is_in(org_select.value))
        ['ror'].drop_nulls().to_list()
    )
    _openaire_ids = (
        nl_openaire_orgs_df
        .filter(pl.col('ror').is_in(_sel_rors))
        ['openaire_orgs_id'].drop_nulls().to_list()
    )
    _force = refresh_btn.value > 0
    _key = hashlib.sha1(','.join(sorted(_sel_rors)).encode()).hexdigest()[:8]
    _cache = CACHE_DIR / f'openaire_pubs_{_key}.parquet'
    if not is_wasm and not _force and _cache.exists():
        openaire_pubs_df = pl.read_parquet(_cache)
    else:
        _openaire_ids_list = '[' + ', '.join(f"'{i}'" for i in _openaire_ids) + ']' if _openaire_ids else "['']"
        openaire_pubs_df = mo.sql(f"""
        SELECT COUNT(DISTINCT pub.id) AS openaire_pubs_count
        FROM openaire.publications AS pub
        WHERE array_length(list_filter(
            pub.organizations,
            x -> list_contains({_openaire_ids_list}, x.id)
        )) > 0
        """, output=False)
        if not is_wasm:
            CACHE_DIR.mkdir(parents=True, exist_ok=True)
            openaire_pubs_df.write_parquet(_cache)
            (CACHE_DIR / '.last_refreshed').touch()
    return (openaire_pubs_df,)


@app.cell(hide_code=True)
def load_cris_pubs_counts(
    CACHE_DIR,
    hashlib,
    is_wasm,
    mo,
    nl_baseline_df,
    org_select,
    pl,
    refresh_btn,
):
    # Count CRIS publications for selected organisations — fast scan of 2.4 M rows.
    _sel_rors = (
        nl_baseline_df
        .filter(pl.col('full_name').is_in(org_select.value))
        ['ror'].drop_nulls().to_list()
    )
    _force = refresh_btn.value > 0
    _key = hashlib.sha1(','.join(sorted(_sel_rors)).encode()).hexdigest()[:8]
    _cache = CACHE_DIR / f'cris_pubs_{_key}.parquet'
    if not is_wasm and not _force and _cache.exists():
        cris_pubs_df = pl.read_parquet(_cache)
    else:
        _rors_clause = ', '.join(f"'{r}'" for r in _sel_rors) if _sel_rors else "''"
        cris_pubs_df = mo.sql(f"""
        SELECT COUNT(*) AS cris_pubs_count
        FROM cris.publications
        WHERE repository_info.ror IS NOT NULL
          AND repository_info.ror IN ({_rors_clause})
        """, output=False)
        if not is_wasm:
            CACHE_DIR.mkdir(parents=True, exist_ok=True)
            cris_pubs_df.write_parquet(_cache)
            (CACHE_DIR / '.last_refreshed').touch()
    return (cris_pubs_df,)


@app.cell(hide_code=True)
def load_openalex_completeness(CACHE_DIR, is_wasm, mo, pl, refresh_btn):
    # aggregate field-level completeness from first OpenAlex works shard (data_0) via direct parquet read
    # the full works table spans 732 files; querying data_0 only keeps this cell fast
    _force = refresh_btn.value > 0
    _cache = CACHE_DIR / 'openalex_completeness.parquet'
    if not is_wasm and not _force and _cache.exists():
        openalex_compl_df = pl.read_parquet(_cache)
    else:
        _WORKS_URL = 'https://objectstore.surf.nl/cea01a7216d64348b7e51e5f3fc1901d:sprouts/data/openalex/works/data_0.parquet'
        _raw = mo.sql(f"""
        SELECT
            COUNT(*)::BIGINT AS total,
            COUNT(title)::BIGINT AS has_title,
            COUNT(abstract_inverted_index)::BIGINT AS has_abstract,
            COUNT(publication_date)::BIGINT AS has_date,
            COUNT_IF(doi IS NOT NULL AND doi LIKE 'https://doi.org/10.%')::BIGINT AS has_doi,
            COUNT_IF(array_length(list_filter(
                authorships,
                x -> array_length(list_filter(
                    x.institutions, y -> y.ror LIKE 'https://ror.org/%'
                )) > 0
            )) > 0)::BIGINT AS has_ror,
            COUNT_IF(array_length(list_filter(
                authorships,
                x -> x.author.orcid LIKE 'https://orcid.org/%'
            )) > 0)::BIGINT AS has_orcid,
            COUNT_IF(array_length(list_filter(
                authorships,
                x -> x.is_corresponding
                  AND array_length(list_filter(x.institutions, y -> y.country_code = 'NL')) > 0
            )) > 0)::BIGINT AS has_nl_corresponding,
            COUNT_IF(open_access.oa_status IS NOT NULL)::BIGINT AS has_oa_status,
            COUNT_IF(
                primary_location.license IS NOT NULL
                AND lower(primary_location.license) LIKE 'cc-%'
            )::BIGINT AS has_cc_license,
            COUNT_IF(array_length(funders) > 0)::BIGINT AS has_funder,
            COUNT(primary_location.source.issn_l)::BIGINT AS has_issn
        FROM read_parquet('{_WORKS_URL}')
        """, output=False)

        _total = _raw['total'][0]
        openalex_compl_df = pl.DataFrame({
            'field': [
                'Title', 'Abstract', 'Publication Date', 'DOI',
                'ROR (affiliation)', 'ORCID (author)',
                'Dutch corresp. author', 'Open Access status',
                'Creative Commons licence', 'Funder / Grant', 'ISSN',
            ],
            'label': [
                'Title', 'Abstract', 'Date', 'DOI',
                'ROR', 'ORCID',
                'NL corresp.', 'OA status',
                'CC licence', 'Funder', 'ISSN',
            ],
            'has_value': [
                _raw['has_title'][0], _raw['has_abstract'][0], _raw['has_date'][0],
                _raw['has_doi'][0], _raw['has_ror'][0], _raw['has_orcid'][0],
                _raw['has_nl_corresponding'][0], _raw['has_oa_status'][0],
                _raw['has_cc_license'][0], _raw['has_funder'][0], _raw['has_issn'][0],
            ],
            'total': [_total] * 11,
        }).with_columns(
            (pl.col('has_value') * 100.0 / pl.col('total')).round(1).alias('pct'),
            (pl.col('total') - pl.col('has_value')).alias('missing'),
        )
        if not is_wasm:
            CACHE_DIR.mkdir(parents=True, exist_ok=True)
            openalex_compl_df.write_parquet(_cache)
            (CACHE_DIR / '.last_refreshed').touch()
    return (openalex_compl_df,)


@app.cell(hide_code=True)
def load_openaire_completeness(CACHE_DIR, is_wasm, mo, pl, refresh_btn):
    # aggregate field-level completeness from first OpenAIRE publications shard (data_0) via direct parquet read
    _force = refresh_btn.value > 0
    _cache = CACHE_DIR / 'openaire_completeness.parquet'
    if not is_wasm and not _force and _cache.exists():
        openaire_compl_df = pl.read_parquet(_cache)
    else:
        _PUBS_URL = 'https://objectstore.surf.nl/cea01a7216d64348b7e51e5f3fc1901d:sprouts/data/openaire/publications/data_0.parquet'
        _raw = mo.sql(f"""
        SELECT
            COUNT(*)::BIGINT AS total,
            COUNT(mainTitle)::BIGINT AS has_title,
            COUNT(publicationDate)::BIGINT AS has_date,
            COUNT_IF(array_length(list_filter(
                pids, x -> x.scheme = 'doi' AND x.value LIKE '10.%'
            )) > 0)::BIGINT AS has_doi,
            COUNT_IF(array_length(list_filter(
                organizations,
                x -> array_length(list_filter(
                    x.pids, y -> y.scheme = 'ROR' AND y.value LIKE 'https://ror.org/%'
                )) > 0
            )) > 0)::BIGINT AS has_ror,
            COUNT_IF(array_length(list_filter(
                authors,
                x -> x.pid.id.scheme = 'orcid' AND x.pid.id.value IS NOT NULL
            )) > 0)::BIGINT AS has_orcid,
            COUNT_IF(bestAccessRight.code IS NOT NULL)::BIGINT AS has_oa_status,
            COUNT_IF(array_length(list_filter(
                instances,
                x -> x.license IS NOT NULL AND (
                    lower(x.license) LIKE '%creativecommons%'
                    OR lower(x.license) LIKE 'cc-%'
                    OR lower(x.license) LIKE 'cc by%'
                )
            )) > 0)::BIGINT AS has_cc_license,
            COUNT_IF(array_length(projects) > 0)::BIGINT AS has_funder,
            COUNT_IF(
                container.issnPrinted IS NOT NULL OR container.issnOnline IS NOT NULL
            )::BIGINT AS has_issn
        FROM read_parquet('{_PUBS_URL}')
        """, output=False)

        _total = _raw['total'][0]
        openaire_compl_df = pl.DataFrame({
            'field': [
                'Title', 'Publication Date', 'DOI',
                'ROR (organisation)', 'ORCID (author)',
                'Open Access status', 'Creative Commons licence',
                'Funder / Grant', 'ISSN',
            ],
            'label': [
                'Title', 'Date', 'DOI',
                'ROR', 'ORCID',
                'OA status', 'CC licence',
                'Funder', 'ISSN',
            ],
            'has_value': [
                _raw['has_title'][0], _raw['has_date'][0], _raw['has_doi'][0],
                _raw['has_ror'][0], _raw['has_orcid'][0],
                _raw['has_oa_status'][0], _raw['has_cc_license'][0],
                _raw['has_funder'][0], _raw['has_issn'][0],
            ],
            'total': [_raw['total'][0]] * 9,
        }).with_columns(
            (pl.col('has_value') * 100.0 / pl.col('total')).round(1).alias('pct'),
            (pl.col('total') - pl.col('has_value')).alias('missing'),
        )
        if not is_wasm:
            CACHE_DIR.mkdir(parents=True, exist_ok=True)
            openaire_compl_df.write_parquet(_cache)
            (CACHE_DIR / '.last_refreshed').touch()
    return (openaire_compl_df,)


@app.cell(hide_code=True)
def load_cris_completeness(CACHE_DIR, is_wasm, mo, pl, refresh_btn):
    # aggregate field-level completeness from full CRIS publications table (2.4 M rows — fast scan)
    _force = refresh_btn.value > 0
    _cache = CACHE_DIR / 'cris_completeness.parquet'
    if not is_wasm and not _force and _cache.exists():
        cris_compl_df = pl.read_parquet(_cache)
    else:
        _raw = mo.sql("""
        SELECT
            COUNT(*)::BIGINT AS total,
            COUNT_IF("cerif:Title" IS NOT NULL AND array_length("cerif:Title") > 0)::BIGINT AS has_title,
            COUNT_IF("cerif:Abstract" IS NOT NULL AND array_length("cerif:Abstract") > 0)::BIGINT AS has_abstract,
            COUNT_IF("cerif:PublicationDate" IS NOT NULL)::BIGINT AS has_date,
            COUNT_IF("cerif:DOI" IS NOT NULL AND "cerif:DOI" LIKE '10.%')::BIGINT AS has_doi,
            COUNT_IF(
                repository_info.ror IS NOT NULL
                AND repository_info.ror LIKE 'https://ror.org/%'
            )::BIGINT AS has_ror,
            COUNT_IF("ar:Access" IS NOT NULL)::BIGINT AS has_oa_status,
            COUNT_IF("cerif:ISSN" IS NOT NULL AND array_length("cerif:ISSN") > 0)::BIGINT AS has_issn
        FROM cris.publications
        """, output=False)

        _total = _raw['total'][0]
        cris_compl_df = pl.DataFrame({
            'field': [
                'Title', 'Abstract', 'Publication Date', 'DOI',
                'ROR (repository)', 'Open Access status', 'ISSN',
            ],
            'label': [
                'Title', 'Abstract', 'Date', 'DOI',
                'ROR', 'OA status', 'ISSN',
            ],
            'has_value': [
                _raw['has_title'][0], _raw['has_abstract'][0], _raw['has_date'][0],
                _raw['has_doi'][0], _raw['has_ror'][0],
                _raw['has_oa_status'][0], _raw['has_issn'][0],
            ],
            'total': [_total] * 7,
        }).with_columns(
            (pl.col('has_value') * 100.0 / pl.col('total')).round(1).alias('pct'),
            (pl.col('total') - pl.col('has_value')).alias('missing'),
        )
        if not is_wasm:
            CACHE_DIR.mkdir(parents=True, exist_ok=True)
            cris_compl_df.write_parquet(_cache)
            (CACHE_DIR / '.last_refreshed').touch()
    return (cris_compl_df,)


@app.cell(hide_code=True)
def constants():
    # define static lookup lists and brand colours used across UI and charts
    CERIF_ENTITIES = [
        'Outputs: Publications',
        'Outputs: Products',
        'Core: Organisation units',
        'Core: Persons',
        'Core: Projects',
    ]
    SOURCES = ['OpenAlex', 'OpenAIRE', 'Crossref', 'ORCID', 'ROR', 'DataCite']
    PUB_TYPES = ['All', 'Journal Article', 'Conference Paper', 'Book Chapter', 'Preprint', 'Thesis', 'Report']
    GROUPING_LABELS = {
        'UNL':   'UNL — Research Universities',
        'VH':    'VH — Universities of Applied Sciences',
        'KNAW':  'KNAW — Royal Academy Institutes',
        'NFU':   'NFU — University Medical Centres',
        'NWO':   'NWO — Research Council',
        'NWO-i': 'NWO-i — NWO Institutes',
        'GOV':   'GOV — Government Research Orgs',
        'INDP':  'INDP — Independent Institutes',
        'UN':    'UN — Other / International',
    }
    return CERIF_ENTITIES, GROUPING_LABELS, PUB_TYPES, SOURCES


@app.cell(hide_code=True)
def ui_base_filters(CERIF_ENTITIES, GROUPING_LABELS, PUB_TYPES, SOURCES, mo):
    # create sector/type/source widgets — these must exist before org_select so that
    # org_select options can be filtered reactively by group_select and barcelona_toggle
    group_select = mo.ui.multiselect(
        options=list(GROUPING_LABELS.keys()),
        value=list(GROUPING_LABELS.keys()),
        label=f"{mo.icon('lucide:layers')} Sector Group",
    )
    barcelona_toggle = mo.ui.switch(
        value=False,
        label=f"{mo.icon('lucide:award')} Barcelona Declaration signatories only",
    )
    entity_select = mo.ui.multiselect(
        options=CERIF_ENTITIES,
        value=['Outputs: Publications'],
        label=f"{mo.icon('lucide:boxes')} CERIF Entity",
    )
    source_select = mo.ui.multiselect(
        options=SOURCES,
        value=['OpenAlex'],
        label=f"{mo.icon('lucide:database')} Primary Source",
    )
    pub_type_select = mo.ui.multiselect(
        options=PUB_TYPES,
        value=['All'],
        label=f"{mo.icon('lucide:file-type')} Type",
    )
    return (
        barcelona_toggle,
        entity_select,
        group_select,
        pub_type_select,
        source_select,
    )


@app.cell(hide_code=True)
def ui_org_select(barcelona_toggle, group_select, mo, nl_baseline_df, pl):
    # build org_select options filtered by the current sector group and Barcelona toggle
    _opts = nl_baseline_df
    if group_select.value:
        _opts = _opts.filter(pl.col('grouping').is_in(group_select.value))
    if barcelona_toggle.value:
        _opts = _opts.filter(pl.col('barcelona_signatory'))
    _org_options = sorted(_opts['full_name'].to_list())
    _default = [v for v in ['University of Amsterdam'] if v in _org_options]
    org_select = mo.ui.multiselect(
        options=_org_options,
        value=_default,
        label=f"{mo.icon('lucide:landmark')} Organisation",
    )
    return (org_select,)


@app.cell(hide_code=True)
def selected_org(
    barcelona_toggle,
    group_select,
    nl_baseline_df,
    nl_openalex_orgs_df,
    org_select,
    pl,
):
    # apply sector-group and Barcelona Declaration filters to the baseline
    _filtered = nl_baseline_df
    if group_select.value:
        _filtered = _filtered.filter(pl.col('grouping').is_in(group_select.value))
    if barcelona_toggle.value:
        _filtered = _filtered.filter(pl.col('barcelona_signatory'))
    filtered_baseline = _filtered

    # resolve selected org names → RORs → OpenAlex rows (for KPI cards)
    _sel_rors = (
        nl_baseline_df
        .filter(pl.col('full_name').is_in(org_select.value))
        ['ror'].drop_nulls().to_list()
    )
    sel_org = nl_openalex_orgs_df.filter(pl.col('ror').is_in(_sel_rors))
    return filtered_baseline, sel_org


@app.cell(hide_code=True)
def header(mo):
    # render the dashboard title and subtitle as markdown
    mo.md("""
    # ORI Data Quality Dashboard
    **Metadata completeness, coverage & enrichment for Dutch research organisations**
    *(OpenAlex · OpenAIRE · Crossref · ORCID — Barcelona Declaration on Open Research Information)*
    """)
    return


@app.cell(hide_code=True)
def sidebar(
    barcelona_toggle,
    cache_last_refreshed,
    entity_select,
    group_select,
    mo,
    org_select,
    pub_type_select,
    refresh_btn,
    source_select,
):
    # stack filter dropdowns and mount them in the marimo sidebar
    filters = mo.vstack([
        mo.md("### Filters"),
        mo.md("---"),
        org_select,
        mo.md("---"),
        group_select,
        barcelona_toggle,
        mo.md("---"),
        entity_select,
        source_select,
        pub_type_select,
    ], gap=1, align='end')

    _last_refreshed_label = (
        f"Last refreshed: **{cache_last_refreshed}**"
        if cache_last_refreshed
        else "_Not yet cached — queries ran live on first load_"
    )
    _refresh_section = mo.vstack([
        mo.md("---"),
        mo.md("### Data Cache"),
        mo.callout(
            mo.md(
                "Refresh re-runs all source queries. "
                "OpenAlex and OpenAIRE scans take **30+ minutes**."
            ),
            kind='warn',
        ),
        refresh_btn,
        mo.md(_last_refreshed_label),
    ], gap=1)

    mo.sidebar(mo.vstack([filters, _refresh_section], gap=2), width="350px")
    return


@app.cell(hide_code=True)
def overview(
    cris_pubs_df,
    filtered_baseline,
    mo,
    nl_baseline_df,
    nl_datasources_df,
    nl_endpoint_df,
    nl_openaire_orgs_df,
    openaire_pubs_df,
    openalex_works_df,
    org_select,
    pl,
):
    # build the overview tab: KPI stat cards, baseline org table, and getting-started guide
    _n_baseline    = nl_baseline_df.height
    _n_filtered    = filtered_baseline.height
    _n_barcelona   = nl_baseline_df.filter(pl.col('barcelona_signatory')).height
    _n_oa_orgs     = nl_openaire_orgs_df.height
    _n_ds_links    = nl_endpoint_df.height
    _n_datasources = nl_datasources_df.height

    # Get record counts from sources
    _openalex_works = openalex_works_df['openalex_works_count'][0] if openalex_works_df.height > 0 else 0
    _openaire_pubs = openaire_pubs_df['openaire_pubs_count'][0] if openaire_pubs_df.height > 0 else 0
    _cris_pubs = cris_pubs_df['cris_pubs_count'][0] if cris_pubs_df.height > 0 else 0

    # Build caption showing selected institution(s), truncated if too many
    _sel = org_select.value
    _sel_caption = (
        ', '.join(_sel[:2]) + (f' +{len(_sel)-2} more' if len(_sel) > 2 else '')
        if _sel else '(none selected)'
    )

    _kpis = mo.hstack([
        mo.stat(
            value=f"{_n_filtered}",
            label="NL Research Orgs (filtered)",
            caption=f"{_n_baseline} total in Zenodo baseline",
            bordered=True,
        ),
        mo.stat(
            value=f"{_n_barcelona}",
            label="Barcelona Declaration NL",
            caption="Signatories matched by ROR",
            bordered=True,
        ),
        mo.stat(
            value=f"{_openalex_works:,}",
            label="OpenAlex Works (est.)",
            caption=_sel_caption,
            bordered=True,
        ),
        mo.stat(
            value=f"{_openaire_pubs:,}",
            label="OpenAIRE Publications",
            caption=_sel_caption,
            bordered=True,
        ),
        mo.stat(
            value=f"{_cris_pubs:,}",
            label="CRIS Publications",
            caption=_sel_caption,
            bordered=True,
        ),
        mo.stat(
            value=f"{_n_ds_links:,}",
            label="Datasource / CRIS Links",
            caption=f"{_n_datasources} unique datasources in OpenAIRE",
            bordered=True,
        ),
        mo.stat(
            value="7",
            label="Sources monitored",
            caption="OpenAlex, OpenAIRE, Crossref, ORCID, ROR, DataCite, CRIS",
            bordered=True,
        ),
    ], gap=3, wrap=True)

    _SECTOR_FULL = {
        'UNL': 'Research Universities', 'VH': 'Univ. Applied Sciences',
        'KNAW': 'Royal Academy Institutes', 'NFU': 'Univ. Medical Centres',
        'NWO': 'Research Council', 'NWO-i': 'NWO Institutes',
        'GOV': 'Government Research', 'INDP': 'Independent Institutes',
        'UN': 'Other / International',
    }
    _org_tbl_data = filtered_baseline.with_columns(
        pl.col('grouping').replace(_SECTOR_FULL).alias('Sector'),
        pl.when(pl.col('barcelona_signatory')).then(pl.lit('✓')).otherwise(pl.lit('')).alias('Barcelona'),
    ).select([
        pl.col('full_name').alias('Organisation'),
        pl.col('acronym').alias('Acronym'),
        pl.col('Sector'),
        pl.col('ror').alias('ROR'),
        pl.col('Barcelona'),
    ])

    _guide = mo.md("""
    ## Getting started

    Use the **filters** (left panel) to focus on a specific organisation, sector group, or Barcelona Declaration signatories.
    The tabs below explore different aspects of data quality:

    - **Completeness** — which metadata fields are present, and for what share of records
    - **Coverage** — which records appear in which open sources, and where there are gaps
    - **Accuracy** — how correctly formatted and valid the field values are
    - **Enrichment** — opportunities to add missing identifiers from external registries

    Data is drawn from the ORI DuckLake catalog (Parquet snapshots on SURF Object Store).
    *Part of the [PID to Portal](https://communities.surf.nl/en/open-research-information) project.*
    """)

    _overview_content = mo.vstack([
        _kpis,
        mo.md(f"### NL Research Organisations *(source: [Zenodo baseline](https://zenodo.org/records/18957154))* — {_n_filtered} of {_n_baseline} shown"),
        mo.ui.table(_org_tbl_data.to_pandas(), selection=None, page_size=15),
        _guide,
    ], gap=4)
    _overview_content
    return


@app.cell(hide_code=True)
def completeness(
    alt,
    cris_compl_df,
    entity_select,
    mo,
    nl_openalex_orgs_df,
    openaire_compl_df,
    openalex_compl_df,
    org_select,
    pl,
    sel_org,
):
    # build the completeness tab: per-source publication field completeness and institution identifier completeness

    def _make_compl_chart(df, title):
        _bar = (
            alt.Chart(df.to_pandas())
            .mark_bar()
            .encode(
                x=alt.X('pct:Q', title='Completeness %', scale=alt.Scale(domain=[0, 100])),
                y=alt.Y('field:N', sort='-x', title=''),
                color=alt.Color(
                    'pct:Q',
                    scale=alt.Scale(
                        domain=[0, 50, 80, 100],
                        range=['#e74c3c', '#f39c12', '#f39c12', '#2ecc71'],
                    ),
                    legend=None,
                ),
                tooltip=[
                    alt.Tooltip('field:N', title='Field'),
                    alt.Tooltip('pct:Q', title='Completeness %', format='.1f'),
                    alt.Tooltip('has_value:Q', title='Records with field', format=','),
                    alt.Tooltip('missing:Q', title='Missing', format=','),
                ],
            )
            .properties(title=title, height=max(220, len(df) * 28), width='container')
        )
        _text = (
            alt.Chart(df.to_pandas())
            .mark_text(dx=4, align='left')
            .encode(
                x=alt.X('pct:Q'),
                y=alt.Y('field:N', sort='-x'),
                text=alt.Text('pct:Q', format='.1f'),
            )
        )
        return (_bar + _text).configure_view(strokeWidth=0)

    def _make_gap_stats(df):
        return mo.hstack([
            mo.stat(
                value=f"{row['pct']:.0f}%",
                label=row['label'],
                caption=f"{row['missing']:,} missing",
                bordered=True,
            )
            for row in df.sort('pct').head(5).to_dicts()
        ], gap=2, wrap=True)

    # -----------------------------------------------------------------------
    # Per-source publication completeness tabs
    # -----------------------------------------------------------------------
    _pub_tabs = mo.tabs({
        'OpenAlex': mo.vstack([
            mo.md(f'*Fields checked per work — first parquet shard ({openalex_compl_df["total"][0]:,} works). '
                  f'Abstract = inverted-index present; ROR/ORCID checked in authorships array.*'),
            _make_gap_stats(openalex_compl_df),
            _make_compl_chart(
                openalex_compl_df,
                f'OpenAlex works completeness — sample ({openalex_compl_df["total"][0]:,} works)',
            ),
        ], gap=2),
        'OpenAIRE': mo.vstack([
            mo.md(f'*Fields checked per publication — first parquet shard ({openaire_compl_df["total"][0]:,} pubs). '
                  f'ROR checked in `organizations[].pids`; ORCID in `authors[].pid`. '
                  f'Abstract and Dutch corresponding author not available in this schema.*'),
            _make_gap_stats(openaire_compl_df),
            _make_compl_chart(
                openaire_compl_df,
                f'OpenAIRE publications completeness — sample ({openaire_compl_df["total"][0]:,} pubs)',
            ),
        ], gap=2),
        'CRIS': mo.vstack([
            mo.md(f'*Fields checked per publication — full table ({cris_compl_df["total"][0]:,} pubs). '
                  f'ROR is repository-level (institution), not per-author. '
                  f'ORCID and funder metadata not available in CERIF-XML schema.*'),
            _make_gap_stats(cris_compl_df),
            _make_compl_chart(
                cris_compl_df,
                f'CRIS publications completeness — full table ({cris_compl_df["total"][0]:,} pubs)',
            ),
        ], gap=2),
    })

    # -----------------------------------------------------------------------
    # Institutions identifier completeness
    # -----------------------------------------------------------------------
    _id_fields = ['has_ror', 'has_grid', 'has_wikidata', 'has_wikipedia', 'has_homepage']
    _id_labels = ['ROR', 'GRID', 'Wikidata', 'Wikipedia', 'Homepage URL']
    _id_pcts   = [
        round(nl_openalex_orgs_df[f].sum() * 100 / nl_openalex_orgs_df.height, 1)
        for f in _id_fields
    ]
    _inst_compl_df = pl.DataFrame({
        'Identifier': _id_labels,
        'pct':        _id_pcts,
        'count':      [nl_openalex_orgs_df[f].sum() for f in _id_fields],
        'total':      [nl_openalex_orgs_df.height] * 5,
    })
    _inst_bar = (
        alt.Chart(_inst_compl_df.to_pandas())
        .mark_bar()
        .encode(
            x=alt.X('pct:Q', title='% of NL institutions', scale=alt.Scale(domain=[0, 105])),
            y=alt.Y('Identifier:N', sort='-x', title=''),
            color=alt.Color(
                'pct:Q',
                scale=alt.Scale(domain=[0, 50, 80, 100], range=['#e74c3c', '#f39c12', '#f39c12', '#2ecc71']),
                legend=None,
            ),
            tooltip=[
                alt.Tooltip('Identifier:N'),
                alt.Tooltip('pct:Q', format='.1f', title='%'),
                alt.Tooltip('count:Q', title='# institutions'),
                alt.Tooltip('total:Q', title='total NL institutions'),
            ],
        )
        .properties(
            title=f'Institution identifier completeness — {nl_openalex_orgs_df.height} NL institutions (OpenAlex)',
            height=200, width='container',
        )
    )

    # -----------------------------------------------------------------------
    # Selected org detail
    # -----------------------------------------------------------------------
    _sel_fields = []
    for _f, _lbl in zip(_id_fields, _id_labels):
        _val = sel_org[_f].any() if sel_org.height > 0 else False
        _sel_fields.append({'Identifier': _lbl, 'Present': '✓' if _val else '✗'})

    _sel_tbl = mo.ui.table(
        pl.DataFrame(_sel_fields).to_pandas(),
        selection=None,
        label=f"Identifier completeness — {', '.join(org_select.value)}",
    ) if sel_org.height > 0 else mo.md('_(no institution selected)_')

    # -----------------------------------------------------------------------
    # Assemble completeness tab
    # -----------------------------------------------------------------------
    _pub_section = mo.vstack([
        mo.md('### Publications metadata completeness *(per-source, first parquet shard)*'),
        _pub_tabs,
    ], gap=2)

    _inst_section = mo.vstack([
        mo.md('### Organisation identifier completeness *(OpenAlex NL institutions)*'),
        _inst_bar,
        _sel_tbl,
    ], gap=2)

    _completeness_content = mo.vstack([
        _pub_section if any(v.startswith('Outputs') for v in entity_select.value) else _inst_section,
    ], gap=4)

    _completeness_content
    return


@app.cell(hide_code=True)
def coverage(alt, mo, nl_openaire_orgs_df, nl_openalex_orgs_df, pl):
    # build the coverage tab: cross-source institution presence and PID completeness comparison
    # -----------------------------------------------------------------------
    # Join OpenAlex institutions with OpenAIRE organizations by ROR
    # -----------------------------------------------------------------------
    _openalex_nl = nl_openalex_orgs_df.select([
        pl.col('display_name').alias('Institution'),
        pl.col('ror'),
        pl.col('works_count').alias('OpenAlex works'),
        pl.col('has_grid').alias('OpenAlex GRID'),
        pl.col('has_wikidata').alias('OpenAlex Wikidata'),
    ])

    _openaire_nl = nl_openaire_orgs_df.filter(pl.col('ror').is_not_null()).select([
        pl.col('ror'),
        pl.col('legalName').alias('OpenAIRE name'),
        pl.col('isni').is_not_null().alias('OpenAIRE ISNI'),
        pl.col('grid').is_not_null().alias('OpenAIRE GRID'),
        pl.col('wikidata').is_not_null().alias('OpenAIRE Wikidata'),
    ])

    _coverage_df = _openalex_nl.join(_openaire_nl, on='ror', how='left').with_columns(
        pl.col('OpenAIRE name').is_not_null().alias('In OpenAIRE'),
    )

    # -----------------------------------------------------------------------
    # Coverage bar: how many NL universities are in OpenAlex vs OpenAIRE
    # -----------------------------------------------------------------------
    _n_in_openalex  = _coverage_df.height
    _n_in_openaire  = _coverage_df.filter(pl.col('In OpenAIRE')).height
    _n_only_openalex = _n_in_openalex - _n_in_openaire

    _bar_data = pl.DataFrame({
        'Source': ['OpenAlex', 'OpenAIRE'],
        'Count':  [_n_in_openalex, _n_in_openaire],
        'Description': ['NL education institutions', 'Matched via ROR'],
    })

    _cov_bar = (
        alt.Chart(_bar_data.to_pandas())
        .mark_bar(cornerRadius=4)
        .encode(
            x=alt.X('Source:N', title=''),
            y=alt.Y('Count:Q', title='# NL institutions'),
            color=alt.Color('Source:N', scale=alt.Scale(
                domain=['OpenAlex', 'OpenAIRE'],
                range=['#009de0', '#2ecc71'],
            )),
            tooltip=['Source:N', 'Count:Q', 'Description:N'],
        )
        .properties(title='NL institution coverage per source', height=220, width=300)
    )

    # -----------------------------------------------------------------------
    # PID completeness: OpenAlex vs OpenAIRE side-by-side
    # -----------------------------------------------------------------------
    _pid_data = pl.DataFrame({
        'PID': ['ROR', 'GRID', 'Wikidata'],
        'OpenAlex %': [
            round(_coverage_df['has_ror'].sum() * 100 / _coverage_df.height, 1) if 'has_ror' in _coverage_df.columns else 100.0,
            round(nl_openalex_orgs_df['has_grid'].sum() * 100 / nl_openalex_orgs_df.height, 1),
            round(nl_openalex_orgs_df['has_wikidata'].sum() * 100 / nl_openalex_orgs_df.height, 1),
        ],
        'OpenAIRE %': [
            round(_openaire_nl.height * 100 / _openalex_nl.height, 1),
            round(_openaire_nl['OpenAIRE GRID'].sum() * 100 / _openaire_nl.height, 1),
            round(_openaire_nl['OpenAIRE Wikidata'].sum() * 100 / _openaire_nl.height, 1),
        ],
    })

    _pid_long = _pid_data.unpivot(
        index='PID', on=['OpenAlex %', 'OpenAIRE %'], variable_name='Source', value_name='pct'
    )

    _pid_chart = (
        alt.Chart(_pid_long.to_pandas())
        .mark_bar()
        .encode(
            x=alt.X('pct:Q', title='% with identifier', scale=alt.Scale(domain=[0, 105])),
            y=alt.Y('PID:N', title=''),
            color=alt.Color('Source:N', scale=alt.Scale(
                domain=['OpenAlex %', 'OpenAIRE %'],
                range=['#009de0', '#2ecc71'],
            )),
            yOffset='Source:N',
            tooltip=['PID:N', 'Source:N', alt.Tooltip('pct:Q', format='.1f')],
        )
        .properties(title='PID completeness: OpenAlex vs OpenAIRE (NL institutions)', height=200, width='container')
    )

    # -----------------------------------------------------------------------
    # Coverage table
    # -----------------------------------------------------------------------
    _cov_tbl_df = _coverage_df.select([
        'Institution',
        pl.lit('✓').alias('OpenAlex'),
        pl.when(pl.col('In OpenAIRE')).then(pl.lit('✓')).otherwise(pl.lit('✗')).alias('OpenAIRE'),
        pl.when(pl.col('OpenAlex GRID')).then(pl.lit('✓')).otherwise(pl.lit('✗')).alias('GRID'),
        pl.when(pl.col('OpenAlex Wikidata')).then(pl.lit('✓')).otherwise(pl.lit('✗')).alias('Wikidata'),
        pl.col('OpenAlex works'),
    ])

    _cov_stats = mo.hstack([
        mo.stat(value=str(_n_in_openalex), label="In OpenAlex", caption="NL education institutions", bordered=True),
        mo.stat(value=str(_n_in_openaire), label="Matched in OpenAIRE", caption="via ROR linkage", bordered=True),
        mo.stat(value=str(_n_only_openalex), label="Only in OpenAlex", caption="Not found in OpenAIRE", bordered=True),
    ], gap=3)

    _coverage_content = mo.vstack([
        mo.md('### Coverage — NL institutions across open sources'),
        mo.md('*Comparing which organisations appear in OpenAlex and OpenAIRE, matched via ROR.*'),
        _cov_stats,
        mo.hstack([
            _cov_bar,
            _pid_chart,
        ], gap=4),
        mo.md('#### Institution × source coverage matrix'),
        mo.ui.table(_cov_tbl_df.to_pandas(), selection=None, page_size=20),
        mo.callout(
            mo.md("**Tip:** Use the ROR to link the same organisation across sources. "
                  "Organisations appearing only in OpenAlex may need to be registered in OpenAIRE's "
                  "[provide.openaire.eu](https://provide.openaire.eu)."),
            kind='info',
        ),
    ], gap=4)

    _coverage_content
    return


@app.cell(hide_code=True)
def accuracy(alt, mo, nl_openalex_orgs_df, openalex_compl_df, pl):
    # build the accuracy tab: identifier format validation checks (ROR, GRID, Wikidata)
    _ror_df = nl_openalex_orgs_df.select([
        pl.col('display_name'),
        pl.col('ror'),
    ]).with_columns([
        pl.col('ror').str.starts_with('https://ror.org/').alias('ror_format_valid'),
        pl.col('ror').str.len_chars().alias('ror_length'),
    ])

    _ror_valid_pct = round(_ror_df['ror_format_valid'].sum() * 100 / _ror_df.height, 1)

    _acc_data = pl.DataFrame({
        'Check': [
            'ROR format valid (https://ror.org/...)',
            'GRID format present',
            'Wikidata format valid (Q number)',
        ],
        'Pass %': [
            _ror_valid_pct,
            round(nl_openalex_orgs_df['has_grid'].sum() * 100 / nl_openalex_orgs_df.height, 1),
            round(
                nl_openalex_orgs_df.filter(
                    pl.col('has_wikidata') & pl.col('wikidata_id').str.starts_with('Q')
                ).height * 100 / nl_openalex_orgs_df.filter(pl.col('has_wikidata')).height
                if nl_openalex_orgs_df.filter(pl.col('has_wikidata')).height > 0 else 0, 1
            ),
        ],
        'Source': ['OpenAlex NL institutions', 'OpenAlex NL institutions', 'OpenAlex NL institutions'],
    })

    _acc_chart = (
        alt.Chart(_acc_data.to_pandas())
        .mark_bar(cornerRadius=4)
        .encode(
            x=alt.X('Pass %:Q', scale=alt.Scale(domain=[0, 105]), title='Pass %'),
            y=alt.Y('Check:N', sort='-x', title=''),
            color=alt.Color(
                'Pass %:Q',
                scale=alt.Scale(domain=[0, 50, 80, 100], range=['#e74c3c', '#f39c12', '#f39c12', '#2ecc71']),
                legend=None,
            ),
            tooltip=['Check:N', alt.Tooltip('Pass %:Q', format='.1f')],
        )
        .properties(title='Identifier format accuracy checks', height=160, width='container')
    )

    _works_completeness_score = round(
        openalex_compl_df.filter(pl.col('field').is_in(['Title', 'DOI', 'Abstract']))['pct'].mean(), 1
    )

    _acc_kpis = mo.hstack([
        mo.stat(
            value=f"{_ror_valid_pct:.0f}%",
            label="ROR format valid",
            caption="Institutions with correct ROR URI",
            bordered=True,
        ),
        mo.stat(
            value=f"{_works_completeness_score:.0f}%",
            label="Core fields avg (works)",
            caption="Title + DOI + Abstract completeness",
            bordered=True,
        ),
        mo.stat(
            value="⚙ Coming soon",
            label="ORCID checksum validation",
            caption="Requires author-level query",
            bordered=True,
        ),
    ], gap=3)

    _accuracy_content = mo.vstack([
        mo.md('### Accuracy — identifier format & value validation'),
        mo.md('*Accuracy checks whether field values conform to the expected format or registry standard.*'),
        _acc_kpis,
        _acc_chart,
        mo.accordion({
            'About these checks': mo.md("""
    **ROR:** Must match `https://ror.org/` followed by a 9-character alphanumeric identifier.
    Example: `https://ror.org/04dkp9463`

    **ORCID:** Must match pattern `0000-0000-0000-000X` (16 digits with dashes, Luhn checksum).

    **DOI:** Must start with `10.` followed by registrant code and suffix.

    **Wikidata:** Entity IDs should start with `Q` followed by digits.

    Full record-level accuracy validation (DOI resolver, ORCID API check) is planned for a future version.
            """),
        }),
    ], gap=4)

    _accuracy_content
    return


@app.cell(hide_code=True)
def enrichment(mo, nl_openalex_orgs_df, openalex_compl_df, pl):
    # build the enrichment tab: prioritised recommendations to fill metadata gaps
    _missing_pub_fields = openalex_compl_df.filter(pl.col('pct') < 80).sort('pct').to_dicts()
    _missing_inst_fields = [
        f for f, has in [
            ('GRID', nl_openalex_orgs_df['has_grid'].mean() < 0.95),
            ('Wikidata', nl_openalex_orgs_df['has_wikidata'].mean() < 0.95),
            ('Wikipedia', nl_openalex_orgs_df['has_wikipedia'].mean() < 0.95),
            ('Homepage URL', nl_openalex_orgs_df['has_homepage'].mean() < 0.95),
        ] if has
    ]

    INTERVENTIONS = [
        {
            'field': 'DOI',
            'priority': 'High',
            'title': 'Register missing DOIs via Crossref',
            'description': (
                f"{openalex_compl_df.filter(pl.col('field')=='DOI')['missing'][0]:,} works lack a DOI. "
                'Register publications with Crossref to obtain DOIs. '
                'This dramatically improves discoverability and linking across sources.'
            ),
            'effort': 'Medium',
            'impact': 'High',
            'action_url': 'https://www.crossref.org/documentation/',
            'action_label': 'Crossref Docs',
        },
        {
            'field': 'ORCID',
            'priority': 'High',
            'title': 'Link ORCID profiles to publications',
            'description': (
                'Author ORCIDs are missing from many records. '
                'Enable institutional ORCID integration to auto-link author profiles to outputs. '
                'Improves attribution and reporting quality.'
            ),
            'effort': 'Low',
            'impact': 'High',
            'action_url': 'https://orcid.org/organizations/integrators',
            'action_label': 'ORCID Integration Guide',
        },
        {
            'field': 'Funder',
            'priority': 'Medium',
            'title': 'Enrich grant metadata via OpenAIRE',
            'description': (
                f"{openalex_compl_df.filter(pl.col('field')=='Funder / Grant')['missing'][0]:,} works lack funder information. "
                'Use the OpenAIRE API to match publications to funded projects and add grant DOIs.'
            ),
            'effort': 'High',
            'impact': 'Medium',
            'action_url': 'https://api.openaire.eu/',
            'action_label': 'OpenAIRE API',
        },
        {
            'field': 'Creative Commons licence',
            'priority': 'Medium',
            'title': 'Add Creative Commons licence to open access works',
            'description': (
                f"{openalex_compl_df.filter(pl.col('field')=='Creative Commons licence')['missing'][0]:,} works lack a CC licence. "
                'Check Unpaywall and re-harvest OA location data to obtain licence strings.'
            ),
            'effort': 'Low',
            'impact': 'Medium',
            'action_url': 'https://unpaywall.org/',
            'action_label': 'Unpaywall API',
        },
        {
            'field': 'Abstract',
            'priority': 'Low',
            'title': 'Harvest abstracts from publisher APIs',
            'description': (
                f"{openalex_compl_df.filter(pl.col('field')=='Abstract')['missing'][0]:,} works have no abstract. "
                'Enrich via Crossref, Semantic Scholar, or Europe PMC APIs.'
            ),
            'effort': 'Medium',
            'impact': 'Low',
            'action_url': 'https://api.semanticscholar.org/',
            'action_label': 'Semantic Scholar API',
        },
        {
            'field': 'Wikidata',
            'priority': 'Low',
            'title': 'Add Wikidata IDs to institutions',
            'description': (
                f"{nl_openalex_orgs_df.filter(~pl.col('has_wikidata')).height} NL institutions are missing Wikidata IDs. "
                'Wikidata enables linked data connections across multiple knowledge bases.'
            ),
            'effort': 'Low',
            'impact': 'Low',
            'action_url': 'https://www.wikidata.org/',
            'action_label': 'Wikidata',
        },
    ]

    _effort_label = {'High': '🔴 High effort', 'Medium': '🟡 Medium', 'Low': '🟢 Low effort'}

    def _make_item(iv):
        return mo.vstack([
            mo.hstack([
                mo.md(f"**[{iv['priority']}]** {iv['title']}"),
                mo.md(f"*{_effort_label[iv['effort']]}* · impact: **{iv['impact']}**"),
            ], justify='space-between', gap=2),
            mo.md(iv['description']),
            mo.md(f"→ [{iv['action_label']}]({iv['action_url']})"),
        ], gap=1)

    _enrichment_content = mo.vstack([
        mo.md('### Enrichment opportunities'),
        mo.md('*Priority recommendations to improve metadata completeness, ranked by impact.*'),
        mo.hstack([
            mo.stat(value=str(sum(1 for i in INTERVENTIONS if i['priority'] == 'High')),
                    label="High-priority actions", bordered=True),
            mo.stat(value=str(sum(1 for i in INTERVENTIONS if i['priority'] == 'Medium')),
                    label="Medium-priority", bordered=True),
            mo.stat(value=str(sum(1 for i in INTERVENTIONS if i['priority'] == 'Low')),
                    label="Low-priority", bordered=True),
        ], gap=3),
        mo.accordion({iv['title']: _make_item(iv) for iv in INTERVENTIONS}),
        mo.callout(
            mo.md("**Next steps:** Implement the suggested enrichment pipelines and re-check this dashboard. "
                  "Questions? Contact [ori-team@surf.nl](mailto:ori-team@surf.nl)."),
            kind='success',
        ),
    ], gap=4)

    _enrichment_content
    return


@app.cell(hide_code=True)
def divider(mo):
    # render a horizontal rule to visually separate the main content from the footer
    mo.md("""
    ---
    """)
    return


@app.cell(hide_code=True)
def footer(mo, org_select):
    from datetime import date
    mo.md(f"""
    <div style="font-size:.75rem;color:#999;display:flex;justify-content:space-between;flex-wrap:wrap;">
    <span>ORI Quality Dashboard · PID to Portal · SURF ORI team</span>
    <span>Organisation: {', '.join(org_select.value) or '(none)'} · Data: OpenAlex / OpenAIRE via SURF DuckLake · {date.today()}</span>
    </div>
    """)
    return


if __name__ == "__main__":
    app.run()
